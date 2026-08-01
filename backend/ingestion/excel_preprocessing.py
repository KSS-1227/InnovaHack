"""
Excel preprocessing pipeline.

Enterprise Compliance Intelligence Platform

Purpose
-------
Process Excel workbooks into the same format expected by MMKGBuilder.

Output
------
texts  : List[str]
images : List[dict]

Images are currently empty because embedded image extraction
is reserved for V2.

Compatible with:
- .xlsx
- .xlsm
"""

from __future__ import annotations

import os
from typing import Dict, List, Tuple

from openpyxl import load_workbook

from ..utils.base import logger


class ExcelChunking:

    def __init__(
        self,
        excel_path: str,
        working_dir: str,
    ):

        self.excel_path = excel_path
        self.working_dir = working_dir

    # ---------------------------------------------------------
    # Public Entry
    # ---------------------------------------------------------

    async def process(
        self
    ) -> Tuple[List[str], List[Dict]]:

        logger.info("📊 Processing Excel workbook...")

        texts = self._extract_text()

        images = []

        logger.info(
            f"✅ Excel Parsed "
            f"({len(texts)} rows extracted)"
        )

        return texts, images

    # ---------------------------------------------------------
    # Extract workbook contents
    # ---------------------------------------------------------

    def _extract_text(self) -> List[str]:

        workbook = load_workbook(
            self.excel_path,
            data_only=True
        )

        extracted = []

        for worksheet in workbook.worksheets:

            extracted.append(
                f"=== SHEET: {worksheet.title} ==="
            )

            for row in worksheet.iter_rows(values_only=True):

                values = []

                for value in row:

                    if value is None:
                        continue

                    value = str(value).strip()

                    if value:
                        values.append(value)

                if values:

                    extracted.append(
                        " | ".join(values)
                    )

        workbook.close()

        return extracted